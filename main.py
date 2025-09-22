import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import pandas as pd
import openpyxl


def get_materials_list():
    # Lê a lista de materiais do arquivo materials.txt
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        materiais_path = os.path.join(base_dir, "Materials", "materials.txt")
        with open(materiais_path, "r", encoding="utf-8") as f:
            materials = [line.strip() for line in f if line.strip()]
        return materials
    except FileNotFoundError:
        print("The file 'materials.txt' was not found.")
        return []

def setup_webdriver():
    # Configura o WebDriver do Chrome
    try:
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--start-maximized")
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        return driver
    except Exception as e:
        print(f"Error setting up WebDriver: {e}")
        return None

def search_materials_dental_cremer(driver, materials):
    # Esta função busca todos os materiais no site da Dental Cremer
    driver.get("https://www.dentalcremer.com.br/")

    try:
        accept_cookies_btn = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="onetrust-accept-btn-handler"]'))
        )
        accept_cookies_btn.click()
        print("Accepted cookies.")

    except Exception as e:
        print(f"Error accepting cookies: {e}")
        pass

    actions = ActionChains(driver)
    list_of_found_materials = []

    for material in materials:
        search_bar = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="search"]'))
        )
        search_bar.clear()
        search_bar.send_keys(material)
        actions.send_keys(Keys.ENTER).perform()
        time.sleep(3)

        select_items_per_page = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="limiter"]'))
        )

        options = select_items_per_page.find_elements(By.TAG_NAME, "option")
        if options:
            options[-1].click()
            time.sleep(3)

        found_items = driver.find_elements(By.CLASS_NAME, "product-item-info")
        if found_items:
            for item in found_items:
                try:
                    item_name = item.find_element(By.CLASS_NAME, "product-item-link").text
                    item_price = item.find_element(By.CSS_SELECTOR, '[data-price-type="finalPrice"] .price').text

                    if not item_price:
                        print("Not item price")
                        continue
                    item_url = item.find_element(By.CLASS_NAME, "product-item-link").get_attribute("href")
                    searched_material = {
                        "name": item_name,
                        "material": material,
                        "price": item_price,
                        "url": item_url
                    }
                    # Busca por qualquer elemento que contenha a classe 'card-label-text'
                    promotion_badge = None
                    try:
                        promotion_badge = item.find_element(By.CSS_SELECTOR, ".card-label-text")
                        # Se existir um <span class="label-text"> dentro, pega o texto dele
                        label_text_elem = promotion_badge.find_element(By.CSS_SELECTOR, ".label-text")
                        promotion_text = label_text_elem.text
                    except Exception:
                        promotion_text = "No"
                    searched_material["promotion"] = promotion_badge.text if promotion_badge else "No"
                    print(searched_material)
                    list_of_found_materials.append(searched_material)
                    print(list_of_found_materials)
                    continue
                except Exception as e:
                    print(f"Error extracting item details: {e}")
        else:
            print(f"No items found for {material}")
            continue
    print(f"Total materials found: {len(list_of_found_materials)}")
    return list_of_found_materials

def create_excel_file(list_of_found_materials):
    try:
        if not list_of_found_materials:
            print("No materials data to write to Excel.")
            return

        df = pd.DataFrame(list_of_found_materials)
        base_dir = os.path.dirname(os.path.abspath(__file__))
        output_path = os.path.join(base_dir, "Materials", "found_materials_2.xlsx")
        df.to_excel(output_path, index=False)
        print(f"Excel file created at {output_path}")

        return output_path
    
    except Exception as e:
        print(f"Error creating Excel file: {e}")

def fill_excel_file(excel_file_path, list_of_found_materials):
    try:
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active

        for idx, material in enumerate(list_of_found_materials, start=2):
            ws.cell(row=idx, column=1, value=material["name"])
            ws.cell(row=idx, column=2, value=material["material"])
            ws.cell(row=idx, column=3, value=material["price"])
            ws.cell(row=idx, column=4, value=material["url"])
            ws.cell(row=idx, column=5, value=material["promotion"])

        wb.save(excel_file_path)
        print(f"Excel file updated at {excel_file_path}")

    except Exception as e:
        print(f"Error filling Excel file: {e}")

def main():
    driver = setup_webdriver()
    if driver is None:
        return
    
    materials = get_materials_list()
    if not materials:
        driver.quit()
        return

    list_of_found_materials = search_materials_dental_cremer(driver, materials)

    if len(list_of_found_materials) > 0:
        excel_file_path = create_excel_file(list_of_found_materials)
        if excel_file_path:
            fill_excel_file(excel_file_path, list_of_found_materials)
        driver.quit()
        return
    
    else:
        print("No materials found.")
        driver.quit()
        return
    
if __name__ == "__main__":
    main()